"""
Report processing logic - FINAL FIXED VERSION
Fixes applied:
  1. UserName column included in all output sheets (taken from UserName column in source data)
  2. URL / Corp_URL taken from the URL column in RCB source data (not generated from username)
  3. High Growth filter: Previous_12M_USD <= $5,000 AND Current_12M_USD >= $50,000 (~15 clients)
  4. 4 sheets: Growth Comparison, High Growth 5K-50K USD, Summary, Exceptions
  5. Column names match expected: CorporateID, CompanyName, UserName, URL,
     Previous_12M_USD, Current_12M_USD, Growth_USD, Growth_%
"""

import pandas as pd
from datetime import datetime


def process_growth_report(df_24m, df_12m, output_file):
    """
    Process growth report from 24-month and 12-month DataFrames.

    Args:
        df_24m: DataFrame with 24-month RCB export data
        df_12m: DataFrame with 12-month RCB export data
        output_file: Output Excel file path (string or Path)

    Returns:
        dict: Summary statistics
    """

    INR_TO_USD = 86  # Exchange rate

    # ──────────────────────────────────────────────────────
    # 1. Prepare 24-month dataframe
    # ──────────────────────────────────────────────────────
    df_24m_prep = df_24m[['CorporateID', 'CorporateName', 'UserName', 'TotalNR1']].copy()
    df_24m_prep.columns = ['CorporateID', 'CorporateName_prev', 'UserName_prev', '24_Month_Revenue']

    # ──────────────────────────────────────────────────────
    # 2. Prepare 12-month dataframe (keep URL if present)
    # ──────────────────────────────────────────────────────
    cols_12m = ['CorporateID', 'CorporateName', 'UserName', 'TotalNR1']
    has_url = 'URL' in df_12m.columns
    if has_url:
        cols_12m.append('URL')
        print("[INFO] URL column found in 12-month data – will use as Corp_URL")
    else:
        print("[INFO] URL column NOT found in 12-month data – will generate from CorporateID")

    df_12m_prep = df_12m[cols_12m].copy()
    new_cols = ['CorporateID', 'CorporateName_curr', 'UserName_curr', '12_Month_Revenue']
    if has_url:
        new_cols.append('URL_curr')
    df_12m_prep.columns = new_cols

    # ──────────────────────────────────────────────────────
    # 3. Merge on CorporateID
    # ──────────────────────────────────────────────────────
    merged = pd.merge(df_24m_prep, df_12m_prep, on='CorporateID', how='outer')
    merged['24_Month_Revenue'].fillna(0, inplace=True)
    merged['12_Month_Revenue'].fillna(0, inplace=True)

    # ──────────────────────────────────────────────────────
    # 4. Calculate metrics
    # ──────────────────────────────────────────────────────
    merged['Previous_12M_Revenue'] = merged['24_Month_Revenue'] - merged['12_Month_Revenue']
    merged['Previous_12M_USD'] = merged['Previous_12M_Revenue'] / INR_TO_USD
    merged['Current_12M_USD']  = merged['12_Month_Revenue']    / INR_TO_USD

    # ──────────────────────────────────────────────────────
    # 5. Separate exceptions (negative USD values)
    # ──────────────────────────────────────────────────────
    exceptions = merged[
        (merged['Previous_12M_USD'] < 0) |
        (merged['Current_12M_USD']  < 0)
    ].copy()

    merged_clean = merged[
        (merged['Previous_12M_USD'] >= 0) &
        (merged['Current_12M_USD']  >= 0)
    ].copy()

    # ──────────────────────────────────────────────────────
    # 6. Growth metrics
    # ──────────────────────────────────────────────────────
    merged_clean['Growth_USD'] = (
        merged_clean['Current_12M_USD'] - merged_clean['Previous_12M_USD']
    )
    merged_clean['Growth_%'] = merged_clean.apply(
        lambda r: (r['Growth_USD'] / r['Previous_12M_USD'] * 100)
                  if r['Previous_12M_USD'] != 0 else 0,
        axis=1
    )

    # ──────────────────────────────────────────────────────
    # 7. UserName: prefer current row value, fall back to previous
    # ──────────────────────────────────────────────────────
    merged_clean['UserName'] = merged_clean['UserName_curr'].fillna(
        merged_clean['UserName_prev']
    )

    # ──────────────────────────────────────────────────────
    # 8. Corp_URL: use URL column from source data; generate fallback
    # ──────────────────────────────────────────────────────
    if 'URL_curr' in merged_clean.columns:
        merged_clean['Corp_URL'] = merged_clean.apply(
            lambda r: r['URL_curr']
            if pd.notna(r['URL_curr']) and str(r['URL_curr']).strip() != ''
            else (
                f"https://rms2.koenig-solutions.com/corporate/{r['CorporateID']}"
                if pd.notna(r['CorporateID']) else ''
            ),
            axis=1
        )
    else:
        merged_clean['Corp_URL'] = merged_clean['CorporateID'].apply(
            lambda cid: f"https://rms2.koenig-solutions.com/corporate/{cid}"
            if pd.notna(cid) and str(cid).strip() != '' else ''
        )

    # ──────────────────────────────────────────────────────
    # 9. Round USD columns to whole numbers
    # ──────────────────────────────────────────────────────
    for col in ['Previous_12M_USD', 'Current_12M_USD', 'Growth_USD']:
        merged_clean[col] = merged_clean[col].round(0).astype(int)

    # ──────────────────────────────────────────────────────
    # 10. Build Growth Comparison sheet
    # ──────────────────────────────────────────────────────
    growth_comparison = merged_clean[[
        'CorporateID', 'CorporateName_curr', 'UserName', 'Corp_URL',
        'Previous_12M_USD', 'Current_12M_USD', 'Growth_USD', 'Growth_%'
    ]].copy()
    growth_comparison.columns = [
        'CorporateID', 'CompanyName', 'UserName', 'URL',
        'Previous_12M_USD', 'Current_12M_USD', 'Growth_USD', 'Growth_%'
    ]
    growth_comparison.sort_values('Growth_USD', ascending=False, inplace=True)
    growth_comparison.reset_index(drop=True, inplace=True)

    print(f"[INFO] Growth Comparison: {len(growth_comparison)} rows")

    # ──────────────────────────────────────────────────────
    # 11. Build High Growth sheet
    #     CORRECT FILTER: Previous <= $5,000 AND Current >= $50,000
    # ──────────────────────────────────────────────────────
    high_growth_mask = (
        (merged_clean['Previous_12M_USD'] <= 5000) &
        (merged_clean['Current_12M_USD']  >= 50000)
    )
    hg_data = merged_clean[high_growth_mask].copy()

    high_growth = pd.DataFrame({
        'CorporateID':      hg_data['CorporateID'].values,
        'CompanyName':      hg_data['CorporateName_curr'].values,
        'UserName':         hg_data['UserName'].values,
        'URL':              hg_data['Corp_URL'].values,
        'Previous_12M_USD': hg_data['Previous_12M_USD'].values,
        'Current_12M_USD':  hg_data['Current_12M_USD'].values,
        'Growth_USD':       hg_data['Growth_USD'].values,
        'Growth_%':         hg_data['Growth_%'].values,
    })
    high_growth.sort_values('Growth_%', ascending=False, inplace=True)
    high_growth.reset_index(drop=True, inplace=True)

    print(f"[INFO] High Growth (Prev<=$5K → Curr>=$50K): {len(high_growth)} clients")
    if len(high_growth) > 0:
        for _, r in high_growth.head(5).iterrows():
            print(f"       {r['CompanyName']:40s}  Prev:${r['Previous_12M_USD']:>8,}  Curr:${r['Current_12M_USD']:>10,}")

    # ──────────────────────────────────────────────────────
    # 12. Build Exceptions sheet
    # ──────────────────────────────────────────────────────
    if len(exceptions) > 0:
        exc_out = exceptions[['CorporateID', 'CorporateName_curr',
                               'Previous_12M_USD', 'Current_12M_USD']].copy()
        exc_out.columns = ['CorporateID', 'CompanyName',
                           'Previous_12M_USD', 'Current_12M_USD']
    else:
        exc_out = pd.DataFrame(columns=['CorporateID', 'CompanyName',
                                         'Previous_12M_USD', 'Current_12M_USD'])

    print(f"[INFO] Exceptions: {len(exc_out)} rows")

    # ──────────────────────────────────────────────────────
    # 13. Build Summary sheet (vertical Metric / Value format)
    # ──────────────────────────────────────────────────────
    top = growth_comparison.iloc[0] if len(growth_comparison) > 0 else None

    def fmt_usd(v):
        return f"${int(v):,}" if v is not None else 'N/A'

    summary_data = {
        'Metric': [
            '★ TOP PERFORMER – BIGGEST MOVER',
            'Company Name',
            'Corporate ID',
            'User Name',
            'Previous 12M Revenue (USD)',
            'Current 12M Revenue (USD)',
            'Growth Amount (USD)',
            'Growth Percentage',
            'Company URL',
            '',
            '■ OVERALL STATISTICS',
            'Total Clients Analyzed',
            'High Growth Clients (Prev ≤$5K → Curr ≥$50K)',
            'Average Previous 12M Revenue (USD)',
            'Average Current 12M Revenue (USD)',
            'Total Growth (USD)',
            'Average Growth % (All Clients)',
            'Total Exceptions',
            '',
            'Report Generated',
        ],
        'Value': [
            '',
            top['CompanyName']  if top is not None else 'N/A',
            str(top['CorporateID']) if top is not None else 'N/A',
            top['UserName']     if top is not None else 'N/A',
            fmt_usd(top['Previous_12M_USD']) if top is not None else 'N/A',
            fmt_usd(top['Current_12M_USD'])  if top is not None else 'N/A',
            fmt_usd(top['Growth_USD'])       if top is not None else 'N/A',
            f"{top['Growth_%']:.1f}%"        if top is not None else 'N/A',
            top['URL']          if top is not None else 'N/A',
            '',
            '',
            len(growth_comparison),
            len(high_growth),
            fmt_usd(growth_comparison['Previous_12M_USD'].mean()),
            fmt_usd(growth_comparison['Current_12M_USD'].mean()),
            fmt_usd(growth_comparison['Growth_USD'].sum()),
            f"{growth_comparison['Growth_%'].mean():.1f}%",
            len(exc_out),
            '',
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        ],
    }
    summary = pd.DataFrame(summary_data)

    # ──────────────────────────────────────────────────────
    # 14. Write Excel with 4 sheets
    # ──────────────────────────────────────────────────────
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        growth_comparison.to_excel(writer, sheet_name='Growth Comparison',    index=False)
        high_growth.to_excel(      writer, sheet_name='High Growth 5K-50K USD', index=False)
        summary.to_excel(          writer, sheet_name='Summary',               index=False)
        exc_out.to_excel(          writer, sheet_name='Exceptions',            index=False)

        # ── Style the Summary sheet ──────────────────────────────────────
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = writer.book
        ws = wb['Summary']
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 55

        gold  = PatternFill('solid', fgColor='FFD700')
        blue  = PatternFill('solid', fgColor='E3F2FD')
        green = PatternFill('solid', fgColor='C8E6C9')

        for row_idx in range(1, ws.max_row + 1):
            val = str(ws.cell(row_idx, 1).value or '')
            if 'TOP PERFORMER' in val or '★' in val:
                for cell in ws[row_idx]:
                    cell.fill = gold
                    cell.font = Font(bold=True, size=13)
                    cell.alignment = Alignment(horizontal='center')
            elif 'OVERALL STATISTICS' in val or '■' in val:
                for cell in ws[row_idx]:
                    cell.fill = green
                    cell.font = Font(bold=True, size=12)
                    cell.alignment = Alignment(horizontal='center')
            elif row_idx in range(3, 11):    # top performer details rows
                for cell in ws[row_idx]:
                    cell.fill = blue
                    if cell.column == 1:
                        cell.font = Font(bold=True, size=11)

    print(f"\n[SUCCESS] Report saved: {output_file}")
    if top is not None:
        print(f"  Top Performer : {top['CompanyName']}")
        print(f"  Growth        : ${top['Growth_USD']:,} ({top['Growth_%']:.1f}%)")

    return {
        'total_clients':       len(growth_comparison),
        'high_growth_clients': len(high_growth),
        'exceptions':          len(exc_out),
        'total_growth_usd':    int(growth_comparison['Growth_USD'].sum()),
        'avg_growth_pct':      round(growth_comparison['Growth_%'].mean(), 1),
        'top_performer':       top['CompanyName']  if top is not None else 'N/A',
        'top_performer_growth':int(top['Growth_USD']) if top is not None else 0,
    }
